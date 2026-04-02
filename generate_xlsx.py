#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
杭州周末情侣约会方案 - Excel 版本
生成时间：2026-03-11 23:45
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# 创建工作簿
wb = Workbook()
wb.title = "杭州周末情侣约会方案"

# 删除默认 sheet
default_sheet = wb.active
wb.remove(default_sheet)

# ============ Sheet 1: 方案总览 ============
ws_overview = wb.create_sheet("方案总览")

# 标题
ws_overview.merge_cells('A1:E1')
ws_overview['A1'] = "💕 杭州周末情侣约会方案（3 月 14-15 日）"
ws_overview['A1'].font = Font(size=18, bold=True, color="FFFFFF")
ws_overview['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_overview['A1'].fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")

# 基本信息
ws_overview['A3'] = "预算上限："
ws_overview['B3'] = "1000 元"
ws_overview['C3'] = "交通方式："
ws_overview['D3'] = "公共交通"
ws_overview['E3'] = "人数："
ws_overview['F3'] = "2 人"

# 表头
headers = ["方案", "类型", "预算 (元)", "浪漫度", "推荐指数", "核心亮点"]
for col, header in enumerate(headers, 1):
    cell = ws_overview.cell(row=5, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")

# 数据
schemes = [
    ["方案 A", "西湖一日游", "400-500", "⭐⭐⭐⭐", "⭐⭐⭐⭐", "经典浪漫，拍照出片"],
    ["方案 B", "乌镇住宿游", "800-1000", "⭐⭐⭐⭐⭐", "⭐⭐⭐⭐⭐", "古镇夜景，文艺复古"],
    ["方案 C", "温泉放松游", "700-900", "⭐⭐⭐⭐", "⭐⭐⭐⭐", "慵懒治愈，私密性好"],
    ["方案 D", "文艺打卡游", "350-450", "⭐⭐⭐", "⭐⭐⭐", "清新拍照，艺术气息"],
]

for row_idx, scheme in enumerate(schemes, 6):
    for col_idx, value in enumerate(scheme, 1):
        ws_overview.cell(row=row_idx, column=col_idx, value=value)

# 设置列宽
ws_overview.column_dimensions['A'].width = 10
ws_overview.column_dimensions['B'].width = 15
ws_overview.column_dimensions['C'].width = 12
ws_overview.column_dimensions['D'].width = 10
ws_overview.column_dimensions['E'].width = 10
ws_overview.column_dimensions['F'].width = 25

# ============ Sheet 2: 方案 A 西湖一日游 ============
ws_a = wb.create_sheet("方案 A-西湖一日游")

ws_a['A1'] = "🌸 方案 A：西湖经典一日游"
ws_a['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_a['A1'].fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
ws_a.merge_cells('A1:C1')

# 预算分解
ws_a['A3'] = "💰 预算分解"
ws_a['A3'].font = Font(bold=True, size=14)

budget_a_headers = ["项目", "费用 (元)", "备注"]
for col, header in enumerate(budget_a_headers, 1):
    ws_a.cell(row=4, column=col, value=header).font = Font(bold=True)

budget_a_data = [
    ["交通", 30, "地铁 + 公交往返"],
    ["游船", 110, "55 元/人×2"],
    ["雷峰塔", 80, "40 元/人×2"],
    ["午餐", 200, "楼外楼/知味观"],
    ["晚餐", 150, "湖滨银泰附近"],
    ["小食饮品", 50, "奶茶、小吃"],
    ["合计", 620, ""],
]

for row_idx, data in enumerate(budget_a_data, 5):
    for col_idx, value in enumerate(data, 1):
        ws_a.cell(row=row_idx, column=col_idx, value=value)

# 行程安排
ws_a['A13'] = "📍 行程安排"
ws_a['A13'].font = Font(bold=True, size=14)

schedule_a_headers = ["时间段", "活动内容", "备注"]
for col, header in enumerate(schedule_a_headers, 1):
    ws_a.cell(row=14, column=col, value=header).font = Font(bold=True)

schedule_a_data = [
    ["9:00-12:00", "断桥残雪→白堤→平湖秋月", "地铁 1 号线龙翔桥站"],
    ["12:00-14:00", "午餐：楼外楼", "推荐：西湖醋鱼、东坡肉"],
    ["14:00-18:00", "乘船游湖→雷峰塔", "含三潭印月"],
    ["18:00-21:00", "晚餐 + 湖滨银泰逛夜市", "可看音乐喷泉 19:00/20:00"],
]

for row_idx, data in enumerate(schedule_a_data, 15):
    for col_idx, value in enumerate(data, 1):
        ws_a.cell(row=row_idx, column=col_idx, value=value)

ws_a.column_dimensions['A'].width = 15
ws_a.column_dimensions['B'].width = 35
ws_a.column_dimensions['C'].width = 25

# ============ Sheet 3: 方案 B 乌镇住宿游 ============
ws_b = wb.create_sheet("方案 B-乌镇住宿游")

ws_b['A1'] = "🏮 方案 B：乌镇古镇住宿游"
ws_b['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_b['A1'].fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
ws_b.merge_cells('A1:C1')

ws_b['A3'] = "💰 预算分解"
ws_b['A3'].font = Font(bold=True, size=14)

budget_b_headers = ["项目", "费用 (元)", "备注"]
for col, header in enumerate(budget_b_headers, 1):
    ws_b.cell(row=4, column=col, value=header).font = Font(bold=True)

budget_b_data = [
    ["高铁", 60, "杭州东→桐乡 30 元/人×2"],
    ["公交", 10, "K282 路 5 元/人×2"],
    ["门票", 520, "西栅 150 元/人×2 + 东栅 110 元/人×2"],
    ["住宿", 350, "古镇内民宿一晚"],
    ["餐饮", 200, "当地特色菜"],
    ["其他", 50, "小吃、饮品"],
    ["合计", 1190, ""],
]

for row_idx, data in enumerate(budget_b_data, 5):
    for col_idx, value in enumerate(data, 1):
        ws_b.cell(row=row_idx, column=col_idx, value=value)

ws_b['A13'] = "📍 行程安排"
ws_b['A13'].font = Font(bold=True, size=14)

schedule_b_headers = ["时间", "活动内容", "备注"]
for col, header in enumerate(schedule_b_headers, 1):
    ws_b.cell(row=14, column=col, value=header).font = Font(bold=True)

schedule_b_data = [
    ["周六 9:00", "杭州东站乘高铁→桐乡站", "约 20 分钟"],
    ["周六 10:00", "K282 公交→乌镇", "约 30 分钟"],
    ["周六 11:00", "入住西栅内民宿", "可多次出入景区"],
    ["周六下午", "游览西栅：染坊→美术馆→书院", ""],
    ["周六晚上", "看西栅夜景 + 晚餐", "夜景超美！"],
    ["周日早上", "晨雾西栅 + 东栅游览", "6:00-8:00 游客少"],
    ["周日中午", "返程", ""],
]

for row_idx, data in enumerate(schedule_b_data, 15):
    for col_idx, value in enumerate(data, 1):
        ws_b.cell(row=row_idx, column=col_idx, value=value)

ws_b.column_dimensions['A'].width = 15
ws_b.column_dimensions['B'].width = 35
ws_b.column_dimensions['C'].width = 25

# ============ Sheet 4: 方案 C 温泉游 ============
ws_c = wb.create_sheet("方案 C-温泉游")

ws_c['A1'] = "♨️ 方案 C：温泉放松之旅"
ws_c['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_c['A1'].fill = PatternFill(start_color="FF7F50", end_color="FF7F50", fill_type="solid")
ws_c.merge_cells('A1:C1')

ws_c['A3'] = "💰 预算分解"
ws_c['A3'].font = Font(bold=True, size=14)

budget_c_headers = ["项目", "费用 (元)", "备注"]
for col, header in enumerate(budget_c_headers, 1):
    ws_c.cell(row=4, column=col, value=header).font = Font(bold=True)

budget_c_data = [
    ["交通", 100, "公交/打车往返"],
    ["温泉 + 住宿", 550, "双人套餐"],
    ["餐饮", 200, "酒店内或附近"],
    ["其他", 50, "饮品、小食"],
    ["合计", 900, ""],
]

for row_idx, data in enumerate(budget_c_data, 5):
    for col_idx, value in enumerate(data, 1):
        ws_c.cell(row=row_idx, column=col_idx, value=value)

ws_c['A11'] = "🏨 推荐酒店"
ws_c['A11'].font = Font(bold=True, size=14)

hotel_headers = ["酒店名称", "位置", "价格", "特色"]
for col, header in enumerate(hotel_headers, 1):
    ws_c.cell(row=12, column=col, value=header).font = Font(bold=True)

hotel_data = [
    ["湍口众安温泉", "临安区湍口镇", "500-700 元", "天然氡温泉，40+ 泡池"],
    ["云曼温泉酒店", "余杭区良渚", "600-700 元", "热带雨林主题"],
    ["杭州湾温泉", "萧山区", "600 元", "园林式，私密性好"],
    ["西湖私汤酒店", "西湖区", "299 元", "西湖旁 + 私汤 + 榻榻米"],
]

for row_idx, data in enumerate(hotel_data, 13):
    for col_idx, value in enumerate(data, 1):
        ws_c.cell(row=row_idx, column=col_idx, value=value)

ws_c.column_dimensions['A'].width = 20
ws_c.column_dimensions['B'].width = 15
ws_c.column_dimensions['C'].width = 15
ws_c.column_dimensions['D'].width = 25

# ============ Sheet 5: 方案 D 文艺打卡 ============
ws_d = wb.create_sheet("方案 D-文艺打卡")

ws_d['A1'] = "🎨 方案 D：文艺打卡一日游"
ws_d['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_d['A1'].fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")
ws_d.merge_cells('A1:C1')

ws_d['A3'] = "💰 预算分解"
ws_d['A3'].font = Font(bold=True, size=14)

budget_d_headers = ["项目", "费用 (元)", "备注"]
for col, header in enumerate(budget_d_headers, 1):
    ws_d.cell(row=4, column=col, value=header).font = Font(bold=True)

budget_d_data = [
    ["交通", 40, "地铁 + 公交"],
    ["门票", 100, "美术馆/展览"],
    ["午餐", 150, "网红餐厅"],
    ["下午茶", 80, "咖啡馆"],
    ["晚餐", 120, "特色餐厅"],
    ["其他", 50, "小食"],
    ["合计", 540, ""],
]

for row_idx, data in enumerate(budget_d_data, 5):
    for col_idx, value in enumerate(data, 1):
        ws_d.cell(row=row_idx, column=col_idx, value=value)

ws_d['A13'] = "📍 推荐路线"
ws_d['A13'].font = Font(bold=True, size=14)

schedule_d_headers = ["时间段", "活动内容", "备注"]
for col, header in enumerate(schedule_d_headers, 1):
    ws_d.cell(row=14, column=col, value=header).font = Font(bold=True)

schedule_d_data = [
    ["9:30-12:00", "中国美术学院象山校区", "免费，需预约"],
    ["12:00-14:00", "象山艺术公社午餐", "创意餐厅"],
    ["14:00-18:00", "浙江美术馆/天目里", "茑屋书店、艺术展"],
    ["18:00-21:00", "晚餐 + 电影/Livehouse", ""],
]

for row_idx, data in enumerate(schedule_d_data, 15):
    for col_idx, value in enumerate(data, 1):
        ws_d.cell(row=row_idx, column=col_idx, value=value)

ws_d.column_dimensions['A'].width = 15
ws_d.column_dimensions['B'].width = 35
ws_d.column_dimensions['C'].width = 25

# ============ Sheet 6: 新增景点 ============
ws_new = wb.create_sheet("新增景点推荐")

ws_new['A1'] = "🆕 最新搜索发现（23:42 更新）"
ws_new['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_new['A1'].fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
ws_new.merge_cells('A1:C1')

ws_new['A3'] = "🌼 季节限定"
ws_new['A3'].font = Font(bold=True, size=14)

new_spot_headers = ["景点名称", "类型", "位置", "价格", "特色"]
for col, header in enumerate(new_spot_headers, 1):
    ws_new.cell(row=4, column=col, value=header).font = Font(bold=True)

new_spot_data = [
    ["包家淇村油菜花海", "季节限定", "富春江边", "免费", "3 月花期，金黄花海"],
    ["杭州爱情博物馆", "情侣打卡", "市中心", "60-80 元/人", "爱情主题展览"],
    ["情人谷", "自然风光", "郊区", "50 元/人", "峡谷木栈道，电影感"],
    ["杨公堤", "散步景点", "西湖区", "免费", "傍晚人少景美"],
    ["小河直街咖啡馆", "休闲", "老城区", "40-60 元", "安静有格调"],
]

for row_idx, data in enumerate(new_spot_data, 5):
    for col_idx, value in enumerate(data, 1):
        ws_new.cell(row=row_idx, column=col_idx, value=value)

ws_new.column_dimensions['A'].width = 20
ws_new.column_dimensions['B'].width = 12
ws_new.column_dimensions['C'].width = 15
ws_new.column_dimensions['D'].width = 12
ws_new.column_dimensions['E'].width = 25

# ============ Sheet 7: 行前准备 ============
ws_prep = wb.create_sheet("行前准备")

ws_prep['A1'] = "📝 行前准备清单"
ws_prep['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_prep['A1'].fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
ws_prep.merge_cells('A1:B1')

prep_headers = ["事项", "状态", "备注"]
for col, header in enumerate(prep_headers, 1):
    ws_prep.cell(row=3, column=col, value=header).font = Font(bold=True)

prep_data = [
    ["确认天气预报", "☐", "3 月中旬杭州 10-20℃"],
    ["预订住宿", "☐", "如选 B/C 方案"],
    ["购买高铁票", "☐", "如选 B 方案，12306 APP"],
    ["准备舒适鞋子", "☐", "步行较多"],
    ["带充电宝", "☐", "拍照耗电"],
    ["带外套", "☐", "早晚温差大"],
    ["带伞", "☐", "可能有小雨"],
]

for row_idx, data in enumerate(prep_data, 4):
    for col_idx, value in enumerate(data, 1):
        ws_prep.cell(row=row_idx, column=col_idx, value=value)

ws_prep['A12'] = "🚌 杭州交通指南"
ws_prep['A12'].font = Font(bold=True, size=14)

traffic_data = [
    ["地铁", "杭州地铁 APP/支付宝乘车码", ""],
    ["公交", "支付宝搜索杭州公交乘车码", ""],
    ["打车", "滴滴/高德", ""],
    ["高铁", "12306 APP", "杭州东→桐乡 30 元"],
]

for row_idx, data in enumerate(traffic_data, 13):
    for col_idx, value in enumerate(data, 1):
        ws_prep.cell(row=row_idx, column=col_idx, value=value)

ws_prep.column_dimensions['A'].width = 20
ws_prep.column_dimensions['B'].width = 10
ws_prep.column_dimensions['C'].width = 30

# 保存文件
output_path = "/Users/yzt/.openclaw/workspace/杭州周末约会方案.xlsx"
wb.save(output_path)

print(f"✅ Excel 文件已生成：{output_path}")
print(f"📊 共 {len(wb.sheetnames)} 个工作表：{', '.join(wb.sheetnames)}")
